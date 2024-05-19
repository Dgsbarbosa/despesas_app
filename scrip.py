import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import math
import pandas as pd
import datetime
import calendar
import copy
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


meses = {1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril', 5: 'maio', 6: 'junho', 7: 'julho', 8: 'agosto', 9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'}



def ler_excel():
    with pd.ExcelFile("Contas Basicas.xlsx") as xlsx:
        
        
        df_receitas = pd.read_excel(xlsx, "Receitas")
        df_despesas = pd.read_excel(xlsx,"Despesas")

  
    
    return df_receitas, df_despesas

   
def organiza_dados(df):     
    
    
    dados = []
    
    for _, conta in df.iterrows():
        
        if pd.isna(conta["vencimento"]):
            conta["vencimento"] = 0
            
        if conta["tipo"] == "semanal":
            conta["vencimento"] = 0
            
        if pd.isna(conta["tipo"]):    
            conta["tipo"] = "semanal"
            
        if pd.isna(conta["parcelas"]) :
              conta["parcelas"] = 1
              
        dado = {"nome": conta['nome'],"vencimento":int(conta["vencimento"]),"valor":conta["valor"],"tipo":conta["tipo"],"parcelas":int(conta["parcelas"]), "primeira parcela":conta["primeira parcela"]}
        dados.append(dado)
        

    return dados

def calendario_do_ano():
    
    dias_da_semana = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sabado', 'domingo']
    global meses
    
    # global meses
    ano = datetime.datetime.now().year
    
    calendario = calendar.Calendar(firstweekday=6)
    
    dias_do_ano = {}
    
    for mes in range(1,13):
        
        dias = []
        numero_semana = 0
        
        for dia in calendario.itermonthdates(ano, mes):
            
            if dia.year == ano and dia.month == mes:    
                
                dia_da_semana = dias_da_semana[dia.weekday()]
                
                if dia_da_semana == "domingo" or numero_semana == 0:
                    numero_semana += 1
                
                # Obtenha o nÃƒÂºmero da semana dentro do mÃƒÂªs  
               
                data_formatada = {"numero_semana":numero_semana,"dia_text":dia_da_semana, "dia": dia.day ,"mes":meses[mes]}
          
                dias.append(data_formatada)
               
             
        
        dias_do_ano[meses[mes]] = dias
    
    
    
    return dias_do_ano

def verifica_parcela(receitas, despesas):
    
    global meses        
    
    for receita in receitas:
        
        receita["meses"] = []
        
        if math.isnan(receita["parcelas"]) == False:
            
            qtd_parcelas = receita["parcelas"]
            mes_primeira_parcela = receita["primeira parcela"]

            receita["valor"] = round(receita["valor"] / qtd_parcelas,2)
            
            for chave_mes, mes_do_ano in meses.items():
                                
                if mes_do_ano == mes_primeira_parcela:
                    
                    ultima_parcela = int(chave_mes + qtd_parcelas - 1)
                    
                    parcelas_excedentes = "" 
                    if ultima_parcela > 12:
                        
                        parcelas_excedentes = ultima_parcela - 12
                        ultima_parcela = 12
                        
                    meses_parcela = []
                    
                    for mes in range(chave_mes,ultima_parcela + 1):
                        
                        meses_parcela.append(meses[mes])
                    
                    
                    if parcelas_excedentes:
                        meses_parcela.append(f" + {parcelas_excedentes}")
                   
                    receita["meses"] = meses_parcela
                 
            
        else:
            receita["parcelas"] = False
            
            if pd.isna(receita["primeira parcela"]) :
                receita["primeira parcela"] = False
            
            
        
        if receita["primeira parcela"] and not receita["meses"]:
            receita["meses"] = [receita["primeira parcela"]]
            
        if receita["primeira parcela"] and not receita["parcelas"]:
            receita["parcelas"] = 1
    
        
 
    for despesa in despesas:
        
        despesa["meses"] = []
        if math.isnan(despesa["parcelas"]) == False:
            
            qtd_parcelas = despesa["parcelas"]
            mes_primeira_parcela = despesa["primeira parcela"]
            despesa["valor"] = round(despesa["valor"] / qtd_parcelas,2)
            for chave_mes, mes_do_ano in meses.items():
                
                if mes_do_ano == mes_primeira_parcela:
                                  
                    chave_mes = int(chave_mes)
                    ultima_parcela = int(chave_mes + qtd_parcelas - 1)
                    
                    parcelas_excedentes = "" 
                    if ultima_parcela > 12:
                        
                        parcelas_excedentes = ultima_parcela - 12
                        ultima_parcela = 12        
                        
                    meses_parcela = []
                    
                    
                    for mes in range(chave_mes,ultima_parcela + 1):
                        
                        meses_parcela.append(meses[mes])
                    
                    if parcelas_excedentes:
                        meses_parcela.append(f" + {parcelas_excedentes}")
                    despesa["meses"] = meses_parcela
            
        else:
            
            despesa["parcelas"] = False
            
            if pd.isna(despesa["primeira parcela"]) :
                despesa["primeira parcela"] = False
            
            
        
        if despesa["primeira parcela"] and not despesa["meses"]:
            despesa["meses"] = [despesa["primeira parcela"]]
            
    return receitas, despesas

def verifica_contas_semanal(lista_contas):
    contas= []
    
    for conta in lista_contas:
        vencimento = conta["vencimento"]
        tipo = conta["tipo"]
        
        
        if vencimento == 0 or tipo == "semanal" or pd.isna(tipo):
            
            contas.append(conta)
    
    return contas          

def verifica_vencimentos(mes):
    
    

    contas = verifica_parcela()
    dias_do_ano = calendario_do_ano()
    
    dias_do_mes = dias_do_ano[mes] 
    
    receitas = contas["receitas"]
    despesas = contas["despesas"]

def cria_dict_meses_semanas(calendario):    
  
    dict_meses= {}
    
    # cria uma chave de um dict para cada mes e semanas  
    
    for mes, dias in calendario.items():
            dict_semanas = {}   
          
            
            
            for dia in dias:
                numero_semana = dia["numero_semana"]
                dict_semanas[numero_semana] = ""          
                
            dict_meses[mes] = dict_semanas
            
            for semana in dict_semanas:
                
                list_dias = []
                
                for dia in dias:
                    
                    if dia["numero_semana"] == semana:
                        list_dias.append(dia)
                
                dict_semanas[semana] = list_dias        
                
            

    return dict_meses

def escolher_pasta():
    
    root = tk.Tk()
    root.withdraw()

    path = filedialog.asksaveasfile(initialfile="Despesas",defaultextension="xlsx",filetypes=[("Excel files","*xlsx")])
    
    
    return path.name

def linhas_planilha(receitas, despesas):
    
    calendario = calendario_do_ano()
    
    
    dict_meses = cria_dict_meses_semanas(calendario)    
   
    
    receitas_copy = copy.deepcopy(receitas)
    despesas_copy = copy.deepcopy(despesas)
    
   
    planilha = {}
    
    # percorre o dicionario de meses que estao separados por mes(chave) e semanas()
    for mes, semanas in dict_meses.items():
        planilha[mes] = {}        
       
        # percorre dict de semanas que tem como  como chave o numero da semana(int) e uma lista com dos dias dessa semana separado por dicts
        for numero_semana, dias in semanas.items():
            
            # listas de despesas semanais
            receitas_semanais = verifica_contas_semanal(receitas_copy)
            despesas_semanais = verifica_contas_semanal(despesas_copy)         
            
           
            linhas = []

            for dia in dias:
                dia_semana = dia["dia"]
                
                # print(receitas_semanais,"\n\n",despesas_semanais)                
              
                linha = {}
                
                # print(dict_meses)
                linha["dia"] = f"{dia['dia_text']}, {dia['dia']} de {dia['mes']}"
             
                # percorre a lista de despesa e verifica os venciementos
                for receita in receitas_copy:
                    nome_receita = receita["nome"]
                    vencimento_receita = int(receita["vencimento"])
                    valor_receita = receita["valor"]
                    tipo_receita = receita["tipo"]
                    meses_parcela_receita = receita["meses"]
                    
                    
                    if dia_semana == vencimento_receita:
                        
                        if tipo_receita == "anual"  :
                            if mes in meses_parcela_receita:
                                linha["nome_receita"] = nome_receita
                                linha["valor_receita"] = valor_receita 
                            
                            else:
                                pass
                        elif tipo_receita == "mensal":
                            linha["nome_receita"] = nome_receita
                            linha["valor_receita"] = valor_receita 
                         
                            
                if not "nome_receita" in linha.keys():
                    try:
                        linha["nome_receita"] = receitas_semanais[0]["nome"]
                        linha["valor_receita"] = receitas_semanais[0]["valor"] 
                        receitas_semanais.pop(0)
                    except:
                        linha["nome_receita"]= ""
                        linha["valor_receita"] = ""
                        pass    
                
                # percorre a lista de despesa e verifica os venciementos  
                for despesa in despesas_copy:
                    nome_despesa = despesa["nome"]
                    vencimento_despesa = int(despesa["vencimento"])
                    valor_despesa = despesa["valor"]
                    tipo_despesa = despesa["tipo"]
                    meses_parcela_despesa = despesa["meses"]

                    if dia_semana == vencimento_despesa:
                        
                        if tipo_despesa == "anual"  :
                            if mes in meses_parcela_despesa:
                                linha["nome_despesa"] = nome_despesa
                                linha["valor_despesa"] =valor_despesa 
                                
                            else:
                                pass
                        elif tipo_despesa == "mensal":
                            linha["nome_despesa"] = nome_despesa
                            linha["valor_despesa"] =valor_despesa 
                         
                            
                if not "nome_despesa" in linha.keys():
                    try:
                        linha["nome_despesa"] = despesas_semanais[0]["nome"]
                        linha["valor_despesa"] = despesas_semanais[0]["valor"] 
                        despesas_semanais.pop(0)
                    except:
                        linha["nome_despesa"] = ""
                        linha["valor_despesa"] = ""
                        
                        pass   
                    
                    
                linhas.append(linha)
                
            planilha[mes][numero_semana] = linhas
           

   
    return planilha

# verifica se o novo vencimento nÃƒÂ£o tem na lista de contas
def valida_vencimento(vencimento, lista_contas):
    
    for conta in lista_contas:
        if int(conta["vencimento"]) == int(vencimento):
            return False
    return True
                    
    
def elimina_datas_iguais(contas,nome_da_conta):

    opcoes = []
    
    # cria uma lista com os dias de vencimento disponiveis   
    for dia in range(1,30):                 
            dia_valido = True
            for c in contas:
                                        
                if dia == c["vencimento"]:
                    dia_valido = False
            if dia_valido:
                opcoes.append(dia)
                
    for conta in contas:
        
        vencimento = int(conta["vencimento"])
        
        vencimentos_iguais = [d for d in contas if int(d["vencimento"]) == vencimento and int(d["vencimento"]) != 0]
       
                    
                    
        if len(vencimentos_iguais) > 1:
                        
            print(f"\nForam encontrados {len(vencimentos_iguais)} {nome_da_conta.lower()} com vencimento no dia: {int(vencimento)}")
            
            print("\nNecessario trocar as datas...")
            for i, vencimento_igual in enumerate(vencimentos_iguais):
                
                index = i + 1 
                nome = vencimento_igual['nome']
                print(f"{index}- {nome.capitalize()}")
            
            while True:
                escolha = input(f"\nQual você deseja alterar? (escolha entre 1 e {len(vencimentos_iguais)}): ")
                
                if escolha.isdigit() and 0 < int(escolha) <= len(vencimentos_iguais) :
                    break
                else:
                    print("\nEscolha um valor valido") 
                    
                        
            item_escolhido = vencimentos_iguais[int(escolha) - 1]
            
            print(f"\nConta escolhida:\n{escolha}- {item_escolhido['nome'].capitalize()}")
            
            
            check = False
                     
                            
                            
            while check == False:
                                 
                
                print("\nOpções: ",end=" ")
                for opcao in opcoes:
                   
                    print(opcao,end=" - ")
                    
                novo_vencimento = input(f"\n\nQual o novo vencimento: ")
                resposta = valida_vencimento(novo_vencimento,contas)
                
                if novo_vencimento.isdigit() and 0 < int(novo_vencimento) < 31:
                    
                    if resposta == True:
                        for el in contas:
                            if el["nome"] == item_escolhido["nome"]:
                                el["vencimento"] = novo_vencimento 
                                check= True
                                opcoes.remove(int(novo_vencimento))
                                
                                print(f"\nA conta {el['nome'].capitalize()} foi alterada para o dia: {novo_vencimento}")
                              
                    else:
                        print(f"Ja hÃ¡ um vencimento no dia {novo_vencimento} na lista")    
                    
               
                else:
                    print("\nDigite um numero de 1 a 31")

# cria a planilha
def cria_planilha(linhas):
           
    ano = datetime.datetime.today().year        
    
    wb = Workbook()
    
    ws = wb.active
    
    wb.remove(ws)
    for mes,semanas in linhas.items():
        
        ws = wb.create_sheet(mes.capitalize())
        
        # adiciona os cabeçalhos
        ws.append([f"Semanas {ano}", "Nome da Receita", "Valor da Receita", "Planejamento", "Nome da Despesa", "Valor da Despesa"])
        
        # adiciona uma linha vasilha
        ws.append([])
        
    
        for semana, dias in semanas.items():
            
            ws.append([f"Semana {semana}"])
            
            # print(semana)
            
            for dia in dias:
                
                # print(dia)
                
                
                ws.append([dia["dia"],dia["nome_receita"],dia["valor_receita"],"",dia["nome_despesa"],dia["valor_despesa"]])

            ws.append([])
            
    # Ajuste a largura das colunas
        for coluna in ws.columns:
            max_length = 0
            coluna = [cell for cell in coluna]
            for cell in coluna:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value) + 2
                except:
                    pass
            ajuste_coluna = get_column_letter(coluna[0].column)  # Obtém a letra da coluna
            ws.column_dimensions[ajuste_coluna].width = max_length     
            
    
             
    caminho = escolher_pasta()    
    
    wb.save(caminho)
    
    
def main():
    
    # le o excel
    df_receitas, df_despesas = ler_excel()
    
    # organiza os dados
    dados_receitas = organiza_dados(df_receitas)
    dados_despesas = organiza_dados(df_despesas)   
    
    
    # verifica se o valor ÃƒÂ© parcelado e adiciona os meses da parcela nos dados
    
    receitas, despesas = verifica_parcela(dados_receitas,dados_despesas)
    
    # elimina datas iguais
    # elimina_datas_iguais(receitas,"Receitas")
    # elimina_datas_iguais(despesas,"Despesas")
    
    # cria as linhas da planilha
    linhas = linhas_planilha(receitas,despesas)
    cria_planilha(linhas)
    
    
    # for mes,semanas in planilha.items():
    #     print(mes)
    #     for semana, dias in semanas.items():
    #         print(semana)
    #         for dia in dias:
    #             print(dia)

if __name__ == "__main__":
    main()

# calendario_do_ano()

# verifica_vencimentos("janeiro")    